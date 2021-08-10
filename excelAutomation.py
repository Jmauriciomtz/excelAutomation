import time
import datetime
from datetime import date
#from datetime import datetime,timedelta,date
#import os,imapclient,email,pyzmail,re,imaplib
import os
import email
import mimetypes
import pyzmail
import re
import imaplib
import imapclient
from email.utils import parseaddr
from pathlib import Path
import zipfile
import win32com.client
import codecs
import PyPDF2
import xlsxwriter
import xlwt
from xlwt import Workbook
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from email.header import decode_header
import bs4
from bs4 import BeautifulSoup
import ast
from tkinter import *
import PIL
from PIL import ImageTk
from PIL import Image
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
from tkinter.ttk import Progressbar
from datetime import timedelta
import requests
from io import BytesIO
# from tika import parser
from imapclient import IMAPClient
import pprint
import re #regular expressions
import shutil
from shutil import copy
import openpyxl


#this will help open excel file in individual folders
import xlrd
from xlrd import open_workbook # this will also help open excel file
import itertools
from itertools import chain
global t
today =date.today()
t = today.strftime("%d-%b-%Y") # da-Mmm-YYY
y = today - timedelta(1)
y = y.strftime('%d-%b-%Y')
excel_file = "C:\\Portadas\\hussmann_portada.xlsx" #global excel file
#########first we have to enter the server from our outlook email to start extracting all emails needed
def RunProgram(username,userpassword,inboxexample,todayscurrentdate,yesterdaysdate): #linked the start button to runnning whole thing
    emails = [] #--------------- contain the body of the emails
    subjects =[] #-------------- contains the subject lines from the emails
    subjectsNoSpace =[] #------------- contains the subject lines with no space between letters
    rawMessages = []
    msgs = [] #------------------ contains the whole information from the whole emails
    attach_directory = [] #-------directory for folder path will be placed here
    aaaList = []
    scacList = []
    bultosList = [] 
    global splitList
    splitList = []
    filePath2 = []
    truckNum = []
    invoiceNum = []
    openExcelFiles =[]
    xx = 0
    final_list = []
    customerNumber = []
    trailerNum = [] #------- contains #, davilan
    scacCodes = [] #-------------- contains all the key value scac codes from subject line
    global scacvalue
    poNum = [] # po numbers from the scac list will be added into this list from the subject line
    BOL = [] #all emails that contain caja directa from the range will be stored here also plataforma directa
    subBOL = [] #contains the emails that have MC1 in subjectline
    totEmail = [] #contains nested list of subjectLine and body
    pdfAttachments = []
    excel_sheet_data = []
    tableBody = []
    attachInvoice = [] #list of all the invoces will be appended into this list from the attachments
    attachCashValue = [] #the cash value from the each pdf will be appended into this list
    attachWeight = [] #the  weight values from each pdf will be appended into this list
    attachBundles = [] #the bundles in pdf will be appended into this list
    attachCxName = [] #cx name sthat will go in the excel sheet like KR REF and others


    def extractEmailInformation():
        server = imapclient.IMAPClient('imap-mail.outlook.com', ssl= True, use_uid = True) #creates connection
        server.login(username+"@uscustombroker.com",userpassword)
        server_info = server.select_folder('INBOX' +inboxexample, readonly=True) #it will look for this emails inside the inbox folder
        #UIDs1=server.search(["SINCE","05-Feb-2020","BEFORE","06-Feb-2020","FROM","alejandra.molina@hussmann.com"])
        #UIDs2=server.search(["SINCE","05-Feb-2020","BEFORE","06-Feb-2020","FROM","juan.j.hernandez@hussmann.com"]) #when using since it includes the day specified, while before dos not.
        #UIDs3=server.search(["SINCE","05-Feb-2020","BEFORE","06-Feb-2020","FROM","adolfo.martinez@hussmann.com"])
        #UIDs=[] 
        #UIDs.append(UIDs1)
        #UIDs.append(UIDs2)
        #UIDs.append(UIDs3)
        UIDs = server.search(['SINCE',yesterdaysdate,'BEFORE',todayscurrentdate,'OR','FROM','juan.j.hernandez@hussmann.com','FROM','alejandra.molina@hussmann.com']) #kkep adding ORs when maore emails listed.
        #print(UIDs)
        f = 0
        for i in UIDs:
            rawMessage = server.fetch([i],['BODY[]']) #the list of UIDs will fetch first argument, second argument is the list which tells fetch to download all the body content for specified UIDs emails
            #the pyzmail parser the raw meesages and returns them as PyzMessage object making subject body to and from field accessible to python code
            msg = pyzmail.PyzMessage.factory(rawMessage[i][b'BODY[]']) #msg is looking
            ##### lower we will download the table from the body to check if it will be krack or not.THIS WAS ONLY BEING USED IF EXCEL'S WERE GOING TO BE ACCEPTED #################
            #new_msg = msg.html_part.get_payload().decode(msg.text_part.charset)
            #soup = BeautifulSoup(new_msg, 'html.parser')
            #newest_msg = soup.prettify()
            #cleaner = re.sub('<.*?>',"",newest_msg) #removes all the taks from html code from table inside the body
            #cleanest = re.sub('\s+',"",cleaner) #removes all the spaces from the table found in body with no more tags
            #cleanest = cleaner.strip('\t\n\r') #this lines takes away all the linebreaks and other whitespace characters
            #bodyKR = re.search(r'(\s)+KR(\s)+',cleanest) #this finds the instance where KR is found in the body
            #nice = re.sub('\s+',"",cleanest) #this line takes away all the spaces and all the data is all together
            #nicee = " ".join(cleanest.split()) #this line leaves single white space between words.
            #if bodyKR != None:
                #f += 1
                #print(nicee) #this works but know that all spaces are gone KR will be hard to be identified 
                #print(f)
                #print("----------------------------------------------------------------------------------------------------------------")
                #splitter = cleanest.split
                #tableBody.append(cleanest)

            subjectLine=msg.get_subject() #this method returns the subject line of all emails. subject line will be needed since some info like trailer number and scac code can be found
            checkSubjectTRUCK = re.search(r'TRUCK',str(subjectLine)) #filtering emails that dont have word TRUCK in subject line
            checkSubjectTRANSFER = re.search(r'TRANSFER',str(subjectLine))#filter emails that dont have the word TRANSFER in subject line
            checkSubjectNoNumber = re.search(r'(\d)+',str(subjectLine))#this will filter out all the emails wit subject line that have no numbers
            checkSubjectRE = re.search(r'RE:',str(subjectLine)) #this will filter all the RE emails
            checkSubjectRe = re.search(r'Re:',str(subjectLine)) #this will filter all the Re emials
            checkSubjectRecall = re.search(r'Recall',str(subjectLine)) #
            if checkSubjectTRUCK != None and checkSubjectRE == None and checkSubjectRe == None and checkSubjectRecall == None: # the word TRUCK must be in subject line while the word RE: must not be presented
            #below we will be getting the body from a raw message. Keep in mind that emails can be sent as plaintext or html
                if msg.text_part != None and msg.html_part == None:
                    body = msg.text_part.get_payload().decode(msg.text_part.charset)
                elif msg.text_part == None and msg.html_part != None:
                    html_string = msg.html_part.get_payload().decode(msg.html_part.charset)  #get_payload returns emails body as a value of bytes data type. decode() takes one argument the message's char encoding stores in text or html_part and return string.
                    #now that you have the html as a string value you need to parse throught that html
                    parsed_html = bs4.BeautifulSoup(html_string,'lxml')
                    newText = parsed_html.get_text() #now that object parsed_html is in text format it can be called with get_text(). This is the new non html text 
                    body = re.sub(r'(<!---[\s\S]*-->)','',newText) #this will substitiue white spaces. THis is the new html text that is know readible. Since html will create allot of spaces this code is used here
                elif msg.text_part == None and msg.html_part == None:
                    print("Body consists of other rather than text and html.")
                else:
                    body = msg.text_part.get_payload().decode(msg.text_part.charset)
                #print(msg.text_part.get_payload().decode(msg.text_part.charset))
                #subjects stored in subject list#
                subjects.append(subjectLine)
                #body of email will be appended into emails list#
                emails.append(body)
                #msg from email will be appended into msgs list#
                msgs.append(msg)
  
        #return subjects #this will return the subjects if program needs to filter some of the emails out by subject line
        #return emails
        #return msgs
        #print(subjects) #prints only subjects from range of emails
        #print("-----------------------------------------------------------------------------")
        #print(emails) #returns only the body of the emails
        #print("------------------------------------------------------------------------------")
        #print(msgs) #msgs stores single emails from range into objects and stores them in certain location
        #return statements dont work well with the print functions
        #print(tableBody) when this string is appended into a list it comes out looking ugly
        print(range(len(msgs)))
        print(range(len(emails)))
        print(range(len(subjects)))
        totEmail.append(msgs)
    extractEmailInformation()
    ############################################################
    def findScacTrailer():
        transfer = {'RODRIGUEZ':'RTVO','RODFRIGUEZ':'RTVO','FERNANDEZ':'FDZL','INDIOCARRIERS':'ICRA','INDIOSCARRIER':'ICRA','RAM':'RMKB','ALA':'ALIA','FEMA':'TFSQ','RAMBIDELEAL':'TRKM','ORTA':'TOSN','LARMONT':'RTUP','ACG':'TNAD','18EXPRESS':'FLPO','MC1':'TAPB','FREIGHTSOL':'FSDG','EXPRESSHORMIGA':'EHSM','TEXASINTERNATIONAL':'TXIE','JIT':'ICRO','TENL':'TEDF','CORSA':'VHAE','SCAC':'JBER','OLYMPIC':'TPYM','OLYMPI':'TPYM','FALYIS':'FYTP','JOT':'PENDING','HIGHVALLEY':'HIGHVALLEY','PROLOG':'PROLOG','REGIOTRANSPORTES':'REGIOTRANSPORTES'}
        transferlist=['RODRIGUEZ','RODFRIGUEZ','FERNANDEZ','INDIOCARRIERS','INDIOSCARRIER','RAM','ALA','FEMA','RAMBIDELEAL','ORTA','LARMONT','ACG','18EXPRESS','MC1','FREIGHTSOL','EXPRESSHORMIGA','TEXASINTERNATIONAL','JIT','TENL','CORSA','SCAC','OLYMPIC','OLYMPI','FALYIS','JOT','HIGHVALLEY','PROLOG','REGIOTRANSPORTES']
        #looks for SCAC code and trailer number in subject line
        #len function returns the number of items in an object. Thus a list is an object
        #range generates a list of numbers used to iterate over for loops.
        for i in range(len(subjects)):
            x = re.sub('\s+',"",str(subjects[i])) #replaces all the white space from the subject line
            subjectsNoSpace.append(x) #appends the subjects lines from picked ranged w/out spaces. 
            print(x) #prints emails subject line w/out spaces not in list for but row form
            #### re's that will pick up trailer number after the keywords listed below ##########
            y = re.search(r'(?<=TRUCK#)(\d)*\d',x) #look after operator, FOCUSES ON NUMBER AFTER TRUCK# ALL NUMBERS
            ah = re.search(r'(?<=TRUCK#)[A-Z]*(\d)*\d',x)
            z = re.search(r'(?<=TRUCK#)[A-Z]*(\d)*\-(\d)*\d',x) #dash, for L, ORTL -----------------------------------------------------------------------------
            zz = re.search(r'(?<=TRUCK#)[A-Z]*(\d)*\d',x) #no dash, for CR,F,FB,H,MTR,ZARO,MFB,NONZ,PE,MSD,GEN,TE,MORFM,EXP,TNI,PE,MVT,D
            zTruck = re.search(r'(?<=TRUCK#)[A-Z]*\-(\d)*-(\d)*\d',x) #double dash for L --------------------------------------------------------
            zzTruck = re.search(r'(?<=TRUCK)[A-Z]*(\d)*\d',x) # no pawn FOR R, CR
            yTruck = re.search(r'(?<=TRUCK)[A-Z]*(\d)*\-(\d)*\d',x) #no pawn -------------------------------------------------------
            TruckFalc = re.search(r'(?<=TRUCKFALCON#)[A-Z]*(\d)*\d',x) #truck falcon all nums
            TruckJB = re.search(r'(?<=TRUCKJ&B#)[A-Z]*(\d)*\d',x) #truck J&B all nums
            yTruckJB = re.search(r'(?<=TRUCKJ&B#)[A-Z]*(\d)*\-(\d)*\d',x) #truck J&B dash and numbs ------------------------------------------
            zTruckJB = re.search(r'(?<=TRUCKJ&B)[A-Z]*(\d)*\d',x) #TRUCK JB NO #
            TruckSta = re.search(r'(?<=TRUCKSTAGECOACH#)[A-Z]*(\d)*\d',x) #truck stagecouch all nums
            zTruckSta = re.search(r'(?<=TRUCKSTAGECOACH#)[A-Z]*(\d)*\-(\d)*\d',x) #truck stagecouch nums and dash --------------------------------
            zTruckRapid = re.search(r'(?<=TRUCKRAPID#)[A-Z]*(\d)*\d',x) #rapid all nums
            TruckFen = re.search(r'(?<=TRUCKFEN#)[A-Z]*(\d)*\-(\d)*\d',x) #fen dash all nums -----------------------------------------------
            zTruckFen = re.search(r'(?<=TRUCKFEN)[A-Z]*(\d)*\d',x)
            zTruckFen2 = re.search(r'(?<=TRUCKFEN#)[A-Z]*(\d)*\d',x)
            TruckFeni = re.search(r'(?<=TRUCKFENI#)[A-Z]*(\d)*-(\d)*\d',x)
            TruckFenix = re.search(r'(?<=TRUCKFENIX#)[A-Z]*(\d)*\d',x) #fenix all nums
            zTruckFenix = re.search(r'(?<=TRUCKFENIX#)[A-Z]*(\d)*\-(\d)*\d',x) # fenix dash and nums -----------------------------------------------
            TruckFenixNo = re.search(r'(?<=TRUCKFENIX)[A-Z]*(\d)*\d',x) # fenix with no #
            zTruckFenixNo = re.search(r'(?<=TRUCKFENIX)[A-Z]*(\d)*\-(\d)*\d',x) #fenix with no # and dash
            TruckAleska = re.search(r'(?<=TRUCKALESKA#)[A-Z]*(\d)*\d',x) #aleska all nums
            TruckLandstar = re.search(r'(?<=TRUCKLANDSTAR#)[A-Z]*(\d)*\d',x) #landstar all nums
            TruckGena = re.search(r'(?<=TRUCKGENA#)[A-Z]*(\d)*\d',x) #gena all nums
            zTruckGena = re.search(r'(?<=TRUCKGENA)[A-Z]*(\d)*\d',x) #gena no pawn letters and nums
            TruckGen = re.search(r'(?<=TRUCKGEN)[A-Z]*(\d)*\d',x)
            zTruckGen = re.search(r'(?<=TRUCKGEN#)[A-Z]*(\d)*\d',x)
            TruckMVT = re.search(r'(?<=TRUCKMVT#)[A-Z]*(\d)*\d',x) #MVT all nums
            TruckJIT = re.search(r'(?<=TRUCKJIT#)[A-Z]*(\d)*\d',x) #JIT all nums
            TruckDial = re.search(r'(?<=TRUCKDIALVAN)[A-Z]*(\d)*\d',x) #Dial no pawn
            zTruckDial = re.search(r'(?<=TRUCKDIALVAN#)[A-Z]*(\d)*\d',x) #dial letters and nums
            TruckCFI = re.search(r'(?<=TRUCKCFI#)[A-Z]*(\d)*\d',x) #truck cfi all nums
            TruckOutWest = re.search(r'(?<=TRUCKOUTWEST#)[A-Z]*(\d)*\d',x) #out west all nums
            TruckPro = re.search(r'(?<=TRUCKPROACTIVE#)[A-Z]*(\d)*\d',x) #proactive letters and nums
            TruckTNI = re.search(r'(?<=TRUCKTNI#)[A-Z]*(\d)*\d',x) #TNI all nums
            TruckBri = re.search(r'(?<=TRUCKBRICKER#)[A-Z]*(\d)*\d',x) #bricker letters and nums
            TruckMill = re.search(r'(?<=TRUCKMILLCREEK#)[A-Z]*(\d)*\d',x) #mill creek 
            TruckMexUs = re.search(r'(?<=TRUCKMEXUSCAN#)[A-Z]*(\d)*\d',x) 
            TruckREG = re.search(r'(?<=TRUCKREG#)[A-Z]*(\d)*\d',x)
            TruckExo = re.search(r'(?<=TRUCKEXO#)[A-Z]*(\d)*\d',x)
            zTruckExo = re.search(r'(?<=TRUCKEXO)[A-Z]*(\d)*\d',x)
            TruckTAP = re.search(r'(?<=TRUCKTAP#)[A-Z]*(\d)*\-(\d)*\d',x) #-----------------------------------------
            zTruckTAP = re.search(r'(?<=TRUCKTAP)[A-Z]*(\d)*\-(\d)*\d',x) #--------------------------------------------
            TruckExp = re.search(r'(?<=TRUCKEXP#)[A-Z]*(\d)*\d',x)
            zTruckExp = re.search(r'(?<=TRUCKEXP)[A-Z]*(\d)*\d',x)
            TruckExOne = re.search(r'(?<=TRUCKEXONE#)[A-Z]*(\d)*\d',x)
            zTruckExOne = re.search(r'(?<=TRUCKEXONE)[A-Z]*(\d)*\d',x)
            TruckMorsa = re.search(r'(?<=TRUCKMORSA#)[A-Z]*(\d)*\d',x)
            gen = re.search(r'(?<=#)[A-Z]*(\d)*\d',x)
           ########################################################################################################################################
            if TruckExOne != None:
                trailerNum.append(TruckExOne.group(0))
            elif zTruckExOne != None:
                trailerNum.append(zTruckExOne.group(0))
            elif TruckMorsa != None:
                trailerNum.append(TruckMorsa.group(0))
            elif TruckExp != None:
                trailerNum.append(TruckExp.group(0))
            elif zTruckExp != None:
                trailerNum.append(zTruckExp.group(0))
            elif TruckTAP != None:
                trailerNum.append(TruckTAP.group(0))
            elif zTruckTAP != None:
                trailerNum.append(zTruckTAP.group(0))
            elif TruckExo != None:
                trailerNum.append(TruckExo.group(0))
            elif zTruckExo != None:
                trailerNum.append(zTruckExo.group(0))
            elif TruckREG != None:
                trailerNum.append(TruckREG.group(0))
            elif TruckMexUs != None:
                trailerNum.append(TruckMexUs.group(0))
            elif TruckMill != None:
                trailerNum.append(TruckMill.group(0))
            elif TruckBri != None:
                trailerNum.append(TruckBri.group(0))
            elif TruckTNI != None:
                trailerNum.append(TruckTNI.group(0))
            elif TruckPro != None:
                trailerNum.append(TruckPro.group(0))
            elif TruckOutWest != None:
                trailerNum.append(TruckOutWest.group(0))
            elif TruckCFI != None:
                trailerNum.append(TruckCFI.group(0))
            elif zTruckDial != None:
                trailerNum.append(zTruckDial.group(0))
            elif TruckDial != None:
                trailerNum.append(TruckDial.group(0))
            elif TruckJIT != None:
                trailerNum.append(TruckJIT.group(0))
            elif TruckMVT != None:
                trailerNum.append(TruckMVT.group(0))
            elif zTruckGen != None:
                trailerNum.append(zTruckGen.group(0))
            elif TruckGen != None:
                trailerNum.append(TruckGen.group(0))
            elif TruckGena != None:
                trailerNum.append(TruckGena.group(0))
            elif zTruckGena != None:
                trailerNum.append(zTruckGena.group(0))
            elif TruckLandstar != None:
                trailerNum.append(TruckLandstar.group(0))
            elif TruckAleska != None:
                trailerNum.append(TruckAleska.group(0))
            elif zTruckFenix != None:
                trailerNum.append(zTruckFenix.group(0))
            elif zTruckFenixNo != None:
                trailerNum.append(zTruckFenixNo.group(0))
            elif TruckFenix != None:
                trailerNum.append(TruckFenix.group(0))
            elif TruckFenixNo != None:
                trailerNum.append(TruckFenixNo.group(0))
            elif TruckFeni != None:
                trailerNum.append(TruckFeni.group(0))
            elif TruckFen != None:
                trailerNum.append(TruckFen.group(0))
            elif zTruckFen2 != None:
                trailerNum.append(zTruckFen2.group(0))
            elif zTruckFen != None:
                trailerNum.append(zTruckFen.group(0))
            elif zTruckRapid != None:
                trailerNum.append(zTruckRapid.group(0))
            elif zTruckSta != None:
                trailerNum.append(zTruckSta.group(0))
            elif TruckSta != None:
                trailerNum.append(TruckSta.group(0))
            elif yTruckJB != None:
                trailerNum.append(yTruckJB.group(0))
            elif TruckJB != None:
                trailerNum.append(TruckJB.group(0))
            elif zTruckJB != None:
                trailerNum.append(zTruckJB.group(0))
            elif TruckFalc != None:
                trailerNum.append(TruckFalc.group(0))
            elif zTruck != None:
                trailerNum.append(zTruck.group(0))
            elif z != None:
                trailerNum.append(z.group(0))
            elif zz != None:
                trailerNum.append(zz.group(0))
            elif yTruck != None:
                trailerNum.append(yTruck.group(0))
            elif zzTruck != None:
                trailerNum.append(zzTruck.group(0))
            elif gen != None:
                trailerNum.append(gen.group(0))
            else:
                print("Not Found")
                print(subjects[i]) # wont be printed out unless an error occurs
            #All the trailer numbers have been isolated and appended into trailerNum
        ################################################################## SCAC code will be used below from subject line ###########################################################
            for w in transferlist:
                my_search = r''+ w #not sure what did those
                y = re.search(my_search,x) #looks for transferlist keywords in subject line
                if y!= None: #if transfer list keywords alocated in subject line it then it enters block of code below
                    #print('Key: '+y.group(0)) #this prints out the whole key and value from trandfer list
                    key = y.group(0)
                    scacvalue = transfer[key]
                    #print(subjects[i])
                    print(key)#key value name of transfer line which will be placed as SCAC value
                    #print('Scac: '+transfer[key]) #this prints out only the key value from subjectline
                    #scacvalue = transfer[key] #holds scac key value from subject line
                    #print(scacvalue) #works
                    scacCodes.append(scacvalue) #list of transfer key have been appended into scacCodes list
        totEmail.append(trailerNum)
        print(scacCodes) #prints list function for scaccodes that are located inside the subject line, but linked to keyvalue that is needed for the insertion of the work sheet
        print(range(len(scacCodes)))
        print(trailerNum) #printa list function for trailerNum must be outside your main for loop and not in between other for loops (works)
        print(range(len(trailerNum)))
    findScacTrailer()
    ################################################################################
    def folderMaker(trailerRef):
        excel_file = "C:\\Portadas\\hussmann_portada.xlsx" #used only for purpouse of attaching excel to main file
        rep1 = filedialog.askdirectory(parent = root, initialdir = "/tmp") #askdirectory will open up file dialog and ask user where to input folder
        for i in trailerRef:
        ### Set newpath to location with all folders
            newpath=(rep1) # this sets the inital path of where info will be imported and distributed
            if not os.path.exists(newpath + "\\" + t): #checks if path exists with t being todays date
                os.makedirs(newpath + "\\"+ t) #creates main folder for range of operation
        ### Get email keyword to make file####
            keyword = i #for amount of trailernums in list, make subfolder of keyword in initial selected file
            if not os.path.exists(newpath+ "\\" + t + '\\'+ keyword):
                    os.makedirs(newpath+ "\\" + t + '\\' + keyword) #creates subfolders depending on amount of non-duplicated emails sent for the range of the operation
            attach_directory.append(newpath + '\\' + t + '\\' + keyword) #attach_directory has the ultimate process for making file. W/out split chipments
        for x in attach_directory:
            copy(excel_file,x) #when files had been ran before make sure you delete this folder because it will fail 
    folderMaker(trailerNum) # trailerNum being used as the argument in call which is interchangable with
    ##################################################################################
    def getAttachments():
    ##### Get attachments ###################
        counter = 1
        ### un packing a MIME message into a directory ###
        for i in range(len(msgs)): #len gives you single value of length of parameter, while range creates a set of numbers from 0 or 1 to len of number
            for part in msgs[i].walk(): #os.walk generates file names in directory tree directory top yields 3 tuples (dirpath, dirnames, filenames).
                if part.get_content_maintype() == "multipart": # multipart are just containers that can be unpacked
                    continue
                filename = part.get_filename() #this breaks down the attachemts by there filename and extension type
                ### write attachments to proper folder ###
                if filename != None and not filename.endswith(".png") and not filename.endswith(".jpg"): #block gets executed when filename not empty and file types are not .png nor .jpg file types
                    precioVenta = re.search(r'Mexicana',filename) #name of mexican filename that is not needed for manifest
                    if precioVenta == None: #anything that is not mexicana type of pdf gets executed
                        try:
                            filePath = os.path.join(attach_directory[i],filename)
                        except:
                            print("New truck line not in list.---------------------------------------------------------------------------------------------------------------------------")
                        with open(filePath,'wb') as f:
                            f.write(part.get_payload(decode=True)) #appends email attachments to emails trailer number into file of choice
    getAttachments()
    ##################################################################################
    
    def findSubInfo():
        y = 0 #counter for subject line
        for i in range(len(subjects)):
            xy = re.sub('\s+',"",str(subjects[i])) #picks up the subject line and removes spaces
            ################# MC1 will be searched for in subjectline and check mark BOL in excel in proper folder name by trailer number ############
            MC1sub = re.search(r'MC1',xy) #looks for MC1 in the subjectline and just. This is just an instance none of that data will be needed.
            if MC1sub != None:
                y += 1
                print(xy) #prints the subject lines that contained the keyword MC1
                subBOL.append(MC1sub) #subBOL contains the instance of MC1 in subject line, but no useful information.
                for trnum in trailerNum:
                    if trnum in xy:
                        print("MC1 was spotted in subject line, excel in file number below will be marked with BOL.-------------------------------------------------------------")
                        print(trnum) #prints the trailer number 
                        wb = openpyxl.load_workbook("C:\\Hussman_manifest\\"+str(t)+"\\"+str(trnum)+"\\hussmann_portada.xlsx")
                        if wb != None:
                            sheet_fill = wb["PORTADA"]
                            sheet_fill["E16"] = "X"
                            wb.save("C:\\Hussman_manifest\\"+str(t)+"\\"+str(trnum)+"\\hussmann_portada.xlsx")
                        else:
                            print("Excel failed to open.")
            
        print(y) #counter for num of times BOL keyword was presented in the subjectline
        #print(subBOL) #prints the list of the amount of times the instance of the keyword was presented
        print(range(len(subBOL))) #this number should match the y counter
    findSubInfo()

    def findBodyInfo(number):
        x = 0 # the counter for the amount of instances keywords are spotted in the body
        for k in range(len(attach_directory)): #for the amount of files created from trailer number
            xz = re.sub('\s+',"",str(emails[k])) #picks up all data from the email
            ##### below were a putting all the casses we will se caja directa, inspecion being found in the body for confirmation BOL #####################
            xCAJADIRECTA = re.search(r'cajadirecta|directa|inspección',xz.lower()) # this covers all the cases making string into lower case
            # when the stirng of words being passed through the value and key word not foudn value will be empty(none/null).
            if xCAJADIRECTA != None:
                print(xCAJADIRECTA.group(0)) #this prints out the case where instance is found in body not the whole body
                if number[k] in attach_directory[k]:
                    print(attach_directory[k]) #prints out the attach directory with the complete path. This will help you see where the BOL is marked
                    files = os.listdir(attach_directory[k])
                    for fl in files:
                        if fl.endswith('.xlsx'):
                            bol_check = openpyxl.load_workbook(attach_directory[k] + '//' + fl)
                            sheet_fill = bol_check["PORTADA"]
                            sheet_fill["E16"] = "X"
                            bol_check.save(attach_directory[k] +'//' + fl)
    findBodyInfo(trailerNum)
    ############### all the information has been attached to its proper folder. Next pdfs will be open to see if there will be any split chipments and make extra folder inside main subfolder ###############################
    def splitMaker():
        ###### if in pdf country: US and Referencia: not Krack (KR) start putting more basic information into this excel sheet #####
        ###### from the ones that are not passing this exceptions will not be used for main excel sheet#################
        ##### once the basic stuff is passed for the pdf it will open the excel and checks if its country: US and its not KRACK ###########
        for f in range(len(attach_directory)): #for lenght of the amount directories inside the list
            files = os.listdir(attach_directory[f]) #list of directories being iretated by the f
            #print(files) 
            for f2 in files:
                if f2.endswith(".pdf"): #works. Filtrs out the excels. 
                    f3 = f2
                    hh = open(attach_directory[f]+"\\"+f3, mode="rb") #opens looping directory with file extensions
                    reader = PyPDF2.PdfFileReader(hh) #sometype of decoder so user can read
                    pages = reader.numPages
                    for num in range(pages): #for length in number of pages. THIS FOR IS RELATED ONLLY TO PDFS
                        page = reader.getPage(num) #gets page while being looped to get hold of all availble pages
                        jk =page.extractText() #extracts information after it has been decoded
                        #print(jk) #correct amount of directories and there pdf information is printed out. All information is in sequence order.
                        #print("-----------------------------------------------------------------------------------------------------------------------")
                        CANADA = re.search(r'CAN(\s)+',jk.upper()) #canada will always come out if in sold to.
                        KR = re.search(r'(?<=:)\s*KR(\s)+',jk.upper()) #this will be used once we have data for this occasion
                        Regular = re.search(r'USA(\s)+',jk.upper()) #this will look for the word USA that is being looked for only for regular entries
                        Mexico = re.search(r'MEX(\s)+',jk.upper()) #this will look for the word MEX in the attachments
                        ############# CANADAS and KRACK CANADAS SUBFOLDER SPLITTER ###############################################
                        if CANADA != None and KR != None:
                            #print(attach_directory[f]+"\\"+f3) #files that will be moved to In_Bond_Krack
                            if not os.path.exists(attach_directory[f]+"\\"+"Krack_In_bond"):
                                os.makedirs(attach_directory[f]+"\\"+"Krack_In_bond") #creates IN bound KRACK folder
                            source = attach_directory[f]+"\\"+f3
                            destination = attach_directory[f]+"\\"+"Krack_In_bond"
                            dest = shutil.copy(source,destination)
                            excel_file = "C:\\Portadas\\hussmann_portada.xlsx"
                            dest = shutil.copy(excel_file,destination)
                        elif CANADA != None:
                            #print(attach_directory[f]+"\\"+f3) #these are the attached files that we want to move into created In_bond file which is made inside there respected trailer number
                            if not os.path.exists(attach_directory[f]+"\\"+"In_bond"):
                                os.makedirs(attach_directory[f]+"\\"+"In_bond") #this creates the file for in_bonds
                            source = attach_directory[f]+"\\"+f3 
                            destination = attach_directory[f]+"\\"+"In_bond"
                            dest = shutil.copy(source,destination) #makes copy of the files related to regular canada into In_bond folder
                            excel_file = "C:\\Portadas\\hussmann_portada.xlsx"
                            dest = shutil.copy(excel_file,destination) #hussmann sheet being placed inside the In_bond folder
                        ################# REGULAR AND KRACK REGULAR SUBFOLDER SPLITTER ##########################################
                        elif Regular != None and KR != None:
                            #print(attach_directory[f]+"\\"+f3) #files that will be moved to Regular entries
                            if not os.path.exists(attach_directory[f]+"\\"+"Krack_Regular"):
                                os.makedirs(attach_directory[f]+"\\"+"Krack_Regular") #creates Regular folder
                            source = attach_directory[f]+"\\"+f3
                            destination = attach_directory[f]+"\\"+"Krack_Regular"
                            dest = shutil.copy(source,destination)
                            excel_file = "C:\\Portadas\\hussmann_portada.xlsx"
                            dest = shutil.copy(excel_file,destination)
                        elif Regular != None:
                            #print(attach_directory[f]+"\\"+f3) #files that will be moved to Regular entries
                            if not os.path.exists(attach_directory[f]+"\\"+"Regular"):
                                os.makedirs(attach_directory[f]+"\\"+"Regular") #creates Regular folder
                            source = attach_directory[f]+"\\"+f3
                            destination = attach_directory[f]+"\\"+"Regular"
                            dest = shutil.copy(source,destination)
                            excel_file = "C:\\Portadas\\hussmann_portada.xlsx"
                            dest = shutil.copy(excel_file,destination)
                        ################# Mexico y Krack Mexico SUBFOLDER SPLITTER ###########################################
                        elif Mexico != None and KR != None:
                            #print(attach_directory[f]+"\\"+f3) #files that will be moved to 
                            if not os.path.exists(attach_directory[f]+"\\"+"Krack_IMP_Mexico"):
                                os.makedirs(attach_directory[f]+"\\"+"Krack_IMP_Mexico") #creates Regular folder
                            source = attach_directory[f]+"\\"+f3
                            destination = attach_directory[f]+"\\"+"Krack_IMP_Mexico"
                            dest = shutil.copy(source,destination)
                            excel_file = "C:\\Portadas\\hussmann_portada.xlsx"
                            dest = shutil.copy(excel_file,destination)
                        elif Mexico != None:
                            #print(attach_directory[f]+"\\"+f3) #files that will be moved to 
                            if not os.path.exists(attach_directory[f]+"\\"+"IMP_Mexico"):
                                os.makedirs(attach_directory[f]+"\\"+"IMP_Mexico") #creates Regular folder
                            source = attach_directory[f]+"\\"+f3
                            destination = attach_directory[f]+"\\"+"IMP_Mexico"
                            dest = shutil.copy(source,destination)
                            excel_file = "C:\\Portadas\\hussmann_portada.xlsx"
                            dest = shutil.copy(excel_file,destination)
            ########### this section will be used if we come back into using excel for the operations ########################
            #for f4 in files:
                #if f4.startswith("hussmann_portada"):
                    #None
                #elif f4.endswith(".xlsx") or f4.endswith(".xls"):
                    #print(attach_directory[f]+"\\"+f4) #prints full path of the directory location of excel files
                    #wb = xlrd.open_workbook(attach_directory[f]+"\\"+f4,'r+') 
    splitMaker()
    ########################## For the lower section program will enter the information of excel that is not needed from the pdfs ##################################################
    def preInjector(): #pre injector will input only the data that is not needed for opening of the attached pdfs. ALL from logic and all ready made lists
        for x in range(len(attach_directory)):
            files = os.listdir(attach_directory[x])
            #print(attach_directory[x]) #prints out the path upto trailer number. X being the trailer number
            for f5 in files:
                #print(attach_directory[x]+"\\"+f5) #prints out path upto files and folders inside the trailer number
                #print("#####################################################################################")
            ####### Below all the spreed sheats witht the subfolder whose name start with Krack will be opened and data will be inputed for customer and customer number #################################
                if f5.startswith("Krack_In_bond") or f5.startswith("Krack_Regular") or f5.startswith("Krack_IMP_Mexico"):
                    #print(attach_directory[x]+"\\"+f5) #prints out the path from krack splits only
                    files2 = os.listdir(attach_directory[x]+"\\"+f5) #list the files from split karck folders only
                    for f6 in files2: #f6 in this for loop represents the files inside the split krack folder 
                        if f6.endswith(".pdf"):
                            None #we dont need the pdfs for this operation only excel sheet and some of the list that we have made all ready
                        elif f6.endswith(".xlsx"):
                            print(attach_directory[x]+"\\"+f5+"\\"+f6) #prints the excel that will have krack as customer
                            loader = openpyxl.load_workbook(attach_directory[x]+"\\"+f5+"\\"+f6)
                            sheet_fill = loader["PORTADA"]
                            sheet_fill["C2"]="KRACK"
                            sheet_fill["C4"]="4201" #KRACK CX #
                            sheet_fill["E10"]="X" #not NAFTA RELATED
                            sheet_fill["C6"]= t  #PRINTS OUT DATE
                            for f in range(len(trailerNum)):
                                #print(trailerNum[f])
                                if str(trailerNum[f]) in str(attach_directory[x]+"\\"+f5+"\\"+f6):
                                    print(trailerNum[f]) #prints the matching trailer number from excel sheet path where that matching excle number will be inputted for trailer number
                                    print(scacCodes[f]) #prints the corresponding scaccode 
                                    sheet_fill["C12"] = str(trailerNum[f])
                                    sheet_fill["C20"] = str(scacCodes[f])
                            loader.save(attach_directory[x]+"\\"+f5+"\\"+f6)
                elif f5.startswith("In_bond") or f5.startswith("Regular") or f5.startswith("IMP_Mexico"):
                    files3 = os.listdir(attach_directory[x]+"\\"+f5) #list the files from non_krack folders only
                    for f7 in files3: #f7 represent the files inside non_krack folders
                        if f7.endswith(".pdf"):
                            None
                        elif f7.endswith(".xlsx"):
                            print(attach_directory[x]+"\\"+f5+"\\"+f7) #prints out excel sheet that will have 
                            abre = openpyxl.load_workbook(attach_directory[x]+"\\"+f5+"\\"+f7)
                            sheet_fill = abre["PORTADA"]
                            sheet_fill["E10"]="x" #this will put x on NAFTA
                            sheet_fill["C6"] = t #prints date
                            for k in range(len(trailerNum)):
                                #print(trailerNum[k])
                                if str(trailerNum[k]) in str(attach_directory[x]+"\\"+f5+"\\"+f7):
                                    print(trailerNum[k]) #prints the matching trailer number from excel sheet path where that matching excle number will be inputted for trailer number
                                    print(scacCodes[k]) #prints the corresponding scaccode 
                                    sheet_fill["C12"] = str(trailerNum[k])
                                    sheet_fill["C20"] = str(scacCodes[k])
                            abre.save(attach_directory[x]+"\\"+f5+"\\"+f7)

    preInjector()
    ########################## For the lower section program will start extracting the needed values from the pdfs and input them into proper excel sheet ##################################################
    def getter():
        pdfkey = ['RI','PU','LC','FW','RE','KR','IS']
        keydic = {'RI':'REACH IN','PU':'PUERTAS','LC':'BOTELLEROS','FW':'FWE','RE':'REFACCIONES','KR':'KRACK','IS':'INSIGHT'}
        for x in range(len(attach_directory)):
            files = os.listdir(attach_directory[x])
            #print(attach_directory[x]) #prints out the path upto trailer number. X being the trailer number
            for f8 in files:
                #print(attach_directory[x]+"\\"+f8) #prints out all files and folders inside the trailer number from todays date folder.
                #print("#####################################################################################")
                if f8.endswith(".pdf") or f8.endswith(".xlsx"):
                    None #this action ignores all the files inside trailernumber. SINCE WE ONLY NEED THE FILES INSIDE SUBFOLDER MADE INSIDE THE TRAILER NUMBER FOLDERR
                else:
                    #print("#####################################################################################")
                    #print(attach_directory[x]+"\\"+f8) #prints out the folder names with preextension isnde the trailernumber folders
                    files4 = os.listdir(attach_directory[x]+"\\"+f8)
                    for f9 in files4:
                        #print(attach_directory[x]+"\\"+f8+"\\"+f9) #prints out the file inside the split folders
                        if f9.endswith(".xlsx"):
                            None #IGNORES THE EXCEL SHET SINCE FOR THIS ACTION WE ARE EXTRACTING THE DATA
                        elif f9.endswith(".pdf"):
                            #print(attach_directory[x]+"\\"+f8+"\\"+f9) #prints out only the pdfs from the split files
                            #### Acess to the pdf files is ready to be opened and extract data #######################
                            kk = open(attach_directory[x]+"\\"+f8+"\\"+f9, mode="rb")
                            fast = PyPDF2.PdfFileReader(kk) #sometype of decoder so user can read
                            paginas = fast.numPages
                            for lnum in range(paginas): #THIS FOR LOOP RELATED ONLY FOR THE PAGES IN THE PDF
                                pagina = fast.getPage(lnum) #gets page while being looped to get hold of all availble pages
                                jkk = pagina.extractText() #extracts information after it has been decoded. All the possble data needed from the pdfs in oder is located in this jkk function
                                #print(jkk) #correct amount of directories and there pdf information is printed out. All information is in sequence order.
                                #print("-----------------------------------------------------------------------------------------------------------------------")
                                remover = re.sub('\s+'," ",jkk) #removes multiple whitespaces between words into single space
                                #print(remover) #data single space. prints all the pdfs data from everysingle file inside split shipment folders
                                invoice = re.search(r'(?<=DATE\s:)\s[\d]+\d',remover)
                                if invoice != None: #appending invoices into list
                                     attachInvoice.append(invoice.group(0))
                                #print(invoice)
                                cashWeight = re.findall(r'(?<=TOTAL)\s[0-9,.]+\s[0-9,.]+',remover) #this section picks up both the total cash value and the total weight value
                                if cashWeight: #appending cash value into list
                                    moneyWeight=cashWeight[0].split(" ") #this splits up cash and weight into diff strings giving the possability to be called by index
                                    mulla = moneyWeight[1]
                                    peso = moneyWeight[2]
                                    attachCashValue.append(mulla)
                                    attachWeight.append(peso)
                                lookBundles = re.search(r'(?<=BULTOS:)\s*\d*\d',remover)
                                if lookBundles != None:
                                    #print(lookBundles.group(0))
                                    attachBundles.append(lookBundles.group(0))
                                getCx = re.search(r'(?<=Referencia:)\s*\w*\w',remover)
                                if getCx != None:
                                    #print(getCx.group(0))
                                    attachCxName.append(getCx.group(0))

        print(attachInvoice)
        print(range(len(attachInvoice)))
        print(attachCashValue)
        print(range(len(attachCashValue)))
        print(attachWeight)
        print(range(len(attachWeight)))
        print(attachBundles)
        print(range(len(attachBundles)))
        print(attachCxName)
        print(range(len(attachCxName)))
    getter()
    ########################### below part will be called the injector which will inject all the data saved in list from previouse def getter #######################################
    def injector():
        cxAbri = ['FROZTEC','RI','PU','LC','FW','RE','AFTERMARKET','KR','TARGET','KRACK-TARGET','MATERIA PRIMA','OTM','GARANTIAS','KC','RMA','INGENIERIA','IS','SGS','INFRINSA']
        cxAbriKey = {'FROZTEC':'4155','RI':'4195','PU':'4196','LC':'4197','FW':'4198','RE':'4199','AFTERMARKET':'4200','KR':'4201','TARGET':'4002','KRACK-TARGET':'4203','MATERIA PRIMA':'4209','OTM':'4215','GARANTIAS':'4253','KC':'4267','RMA':'4300','INGENIERIA':'4306','IS':'6009','SGS':'6010','INFRINSA':'6029'}
        valueLi = []
        weightLi = []
        bundlesLi = []
        for t in range(len(attach_directory)):
            files = os.listdir(attach_directory[t])
            for f10 in files:
                addedValue = 0
                addedWeight = 0
                addedBundles = 0
                lastPdf=[] #list that pdfs will be appended into
                #print(attach_directory[t]+"\\"+f10)
                if f10.endswith(".pdf") or f10.endswith(".xlsx"): #f10 is the trailer number folders
                    None #this filters out the pdf and xl so we only have folders left
                else:
                    #print(attach_directory[t]+"\\"+f10)
                    files3 = os.listdir(attach_directory[t]+"\\"+f10)
                    #print(files3) #prints everything inside the split folders which contain excel and copy of the pdfs
                    for f11 in files3: #f11 is everything everything inside the split folders
                        if f11.endswith(".pdf"): #focuses only on the pdfs so they can be appended and refreshed since list is local
                            lastPdf.append(f11) #appends just the needed items for the directory. Last pdf is only used for the range of invoices
                    load = openpyxl.load_workbook(attach_directory[t]+"\\"+f10+"\\"+f11)
                    activeSheet = load["PORTADA"]
                    # if else block below will be used only for range of invoices in 
                    if lastPdf[0] == lastPdf[-1]:
                        activeSheet["C14"]=lastPdf[0][:-16] #this section is meant split folders that only have one inbox in them
                    else:
                        activeSheet["C14"] = lastPdf[0][:-16]+"-"+lastPdf[-1][4:-16] #this section inputs the invoice number with the dash if they have morte than one invoice in split folder
                    ############### below cx name name and number will be collected from the invoice that is inputed into the excel sheet ##########################################
                    for last in range(len(lastPdf)):
                        faster =  PyPDF2.PdfFileReader(attach_directory[t]+"\\"+f10+"\\"+lastPdf[last]) #this works because its getting called with the appended values from f11 (the split folders pdfs)
                        paginas = faster.numPages
                        for num in range(paginas): #THIS FOR LOOP RELATED ONLY FOR THE PAGES IN THE PDF AT PERIOD IN TIME
                            pagina = faster.getPage(num) #gets page while being looped to get hold of all availble pages
                            jkk = pagina.extractText()
                            ty = re.sub('\s+'," ",jkk) #subs multiple spaces into single spaces
                            #print(jkk)
                            moneyHeavy  = re.findall(r'(?<=TOTAL)\s[0-9,.]+\s[0-9,.]+',ty) #this section picks up both the total cash value and the total weight value
                            if moneyHeavy:
                                letsSplit = moneyHeavy[0].split(" ")
                                mmoneyy = (letsSplit[1]) #holds dollar value from invoice in point of time
                                wweightt = (letsSplit[2]) #holds weight value from invoice in point of time
                                moneyy = re.sub(",","",mmoneyy)
                                weightt = re.sub(",","",wweightt) #SUBS THE COMMAS TO NO SPACE
                                money = float(moneyy)
                                weight = float(weightt)
                                addedValue += money
                                addedWeight += weight
                            ontanBundles = re.search(r'(?<=BULTOS:)\s*\d*\d',ty)
                            if ontanBundles != None:
                                th =int(ontanBundles.group(0))
                                addedBundles += th
                    activeSheet["C8"] = addedValue
                    activeSheet["C10"] = addedWeight
                    activeSheet["C16"] = addedBundles
                    valueLi.append(addedValue)
                    weightLi.append(addedWeight)
                    bundlesLi.append(addedBundles)
                    print(valueLi)
                    print(weightLi)
                    print(bundlesLi)
                    faster =  PyPDF2.PdfFileReader(attach_directory[t]+"\\"+f10+"\\"+lastPdf[0]) #this works because its getting called with the appended values from f11 (the split folders pdfs)
                    paginas = faster.numPages
                    for num in range(paginas): #THIS FOR LOOP RELATED ONLY FOR THE PAGES IN THE PDF AT PERIOD IN TIME
                        pagina = faster.getPage(num) #gets page while being looped to get hold of all availble pages
                        jkk = pagina.extractText()
                        ty = re.sub('\s+'," ",jkk) #subs multiple spaces into single spaces
                        #print(ty) #prints data from all atached files
                        #print("---------------------------------------------------------------------")
                        ### block below will need get cx from pdf and cx # from dictoanry made above ###################
                        cuhstomer = re.search(r'(?<=Referencia:)\s*\w*\s',ty) #returns the cx
                        if cuhstomer != None:
                            print(ty)
                            print(cuhstomer.group(0))
                            print(type(cuhstomer.group(0)))
                            print(len(cuhstomer.group(0)))
                            print("-------------------------------------------------")
                            activeSheet["C2"] = cuhstomer.group(0) # this section will load the work book with the proper cx name
                            if cuhstomer.group(0) == " FROZTEC ": #nested if is becasue of the purpouse outter if has value cuhstomer.group(0)
                                activeSheet["C4"] = "4155"
                            elif cuhstomer.group(0) == " RI ":
                                activeSheet["C4"] = "4195"
                            elif cuhstomer.group(0) == " PU ":
                                activeSheet["C4"] = "4196"
                            elif cuhstomer.group(0) == " LC ":
                                activeSheet["C4"] = "4197"
                            elif cuhstomer.group(0) == " FW ":
                                activeSheet["C4"] = "4198"
                            elif cuhstomer.group(0) == " RE ":
                                activeSheet["C4"] = "4199"
                            elif cuhstomer.group(0) == " AFTERMARKET ":
                                activeSheet["C4"] = "4200"
                            elif cuhstomer.group(0) == " KR ":
                                activeSheet["C4"] = "4201"
                            elif cuhstomer.group(0) == " TARGET ":
                                activeSheet["C4"] = "4002"
                            elif cuhstomer.group(0) == " KRACK-TARGET ":
                                activeSheet["C4"] = "4203"
                            elif cuhstomer.group(0) == " MATERIA PRIMA ":
                                activeSheet["C4"] = "4209"
                            elif cuhstomer.group(0) == " OTM ":
                                activeSheet["C4"] = "4215"
                            elif cuhstomer.group(0) == " GARANTIAS ":
                                activeSheet["C4"] = "4253"
                            elif cuhstomer.group(0) == " KC ":
                                activeSheet["C4"] = "4267"
                            elif cuhstomer.group(0) == " RMA ":
                                activeSheet["C4"] = "4300"
                            elif cuhstomer.group(0) == " INGENIERIA ":
                                activeSheet["C4"] = "4306"
                            elif cuhstomer.group(0) == " IS ":
                                activeSheet["C4"] = "6009"
                            elif cuhstomer.group(0) == " SGS ":
                                activeSheet["C4"] = "6010"
                            elif cuhstomer.group(0) == " INFRINSA ":
                                activeSheet["C4"] = "6029"
                        ###### Below block will be used to decide if sheet will be related or not #########################
                        Mexicoo = re.search(r'MEX(\s)+',ty) #this will look for the word MEX in the attachments
                        if Mexicoo == None:
                            activeSheet["E8"] = " X"
                    load.save(attach_directory[t]+"\\"+f10+"\\"+f11)
    injector()

    def lastMin():
        BOLpath = [] #list where BOL paths upto to trailer number will be appended
        for g in range(len(attach_directory)):
            files = os.listdir(attach_directory[g])
            files2 = os.listdir(attach_directory[g])
            files4 = os.listdir(attach_directory[g])
            #print(files) #prints all the files and folders of trailer number folder in a list
            #print(attach_directory[g]) #prints directory from upto trailer number folder
            ## for loop below will be used to append directories that have BOL in subfolder sheet that will be used to compared with split subfolder sheets ##############
            for f12 in files:
                if f12.startswith("hussmann_portada"):
                    #print(attach_directory[g]+"\\"+f12) #subfolder excel ready to be opend
                    subSheet = openpyxl.load_workbook(attach_directory[g]+"\\"+f12) #trailer folder excel will open with this command
                    subPortada = subSheet.active
                    checkBOL = subPortada["E16"]
                    #print(checkBOL.value)
                    if checkBOL.value == "X":
                        #print(attach_directory[g]+"\\"+f12) #prints the direct path to the subfolder excel sheet that has the X o for the BOL
                        #print(attach_directory[g]) #prints the trailer number path for sub excel sheet that will have the X for the BOL
                        BOLpath.append(attach_directory[g]) #appends the BOL upto trailer number
            ### for loop below will used to input the X in the splitfolder from the list appended above ####
            for f13 in files2:
                if f13.endswith(".pdf") or f13.endswith(".xlsx"):
                    None
                else:
                    files3 = os.listdir(attach_directory[g]+"\\"+f13)
                    for f14 in files3:
                        if f14.endswith(".xlsx"):
                            #print(attach_directory[g]+"\\"+f13+"\\"+f14)
                            if attach_directory[g] not in BOLpath: #disgrads all the paths that are not inside the list of paths which have a BOL
                                None
                            else:
                                ### code below goes into split subfolder and inputs the X in appropiate sheet ###
                                #print(attach_directory[g]+"\\"+f13+"\\"+f14) #prints out the full path upto the excel sheet where x should be marked for BOL
                                semiLastBOL = openpyxl.load_workbook(attach_directory[g]+"\\"+f13+"\\"+f14)
                                subPortadita = semiLastBOL["PORTADA"]
                                subPortadita["E16"] = "X"
                                semiLastBOL.save(attach_directory[g]+"\\"+f13+"\\"+f14)
            ### this section will open when 
            for f15 in files4:
                if f15.endswith(".pdf") or f15.endswith(".xlsx"):
                    None
                else:
                    files5 = os.listdir(attach_directory[g]+"\\"+f15)
                    #print(attach_directory[g]+"\\"+f15)
                    #print(f15)
                    #print(type(f15)) #type is string
                    if f15 == "In_bond" or f15 == "Krack_In_bond": 
                        #print(f15) #prints out only whats asked for
                        for f16 in files5:
                            print(attach_directory[g]+"\\"+f15+"\\"+f16)
                            if f16.endswith(".pdf"):
                                None
                            else:
                                lastOpen = openpyxl.load_workbook(attach_directory[g]+"\\"+f15+"\\"+f16)
                                fill = lastOpen["PORTADA"]
                                fill["E16"] = "X"
                                lastOpen.save(attach_directory[g]+"\\"+f15+"\\"+f16)
                        

    lastMin()

root=tk.Tk() #global
root.title("Email extractor and Data Entry:Hussmann")
username = StringVar()
userpassword = StringVar()
inboxexample = StringVar()
fromdate = StringVar()
currentdate = StringVar()

###Icon for program top left
#root.iconbitmap()
root.configure(bg ='gray63')
canvas = Canvas(root,width = 640, height = 50,bg = "snow")
canvas.grid(row = 12, column = 0,columnspan=6,padx = 5,pady = 5)
canvas.create_text(320,20,font = "Times 14 bold", text="Property of Eduardo Lozano and Company")
url = requests.get("https://uscustombroker.com/site/wp-content/uploads/Logo-16.jpg")
#uopen = urlopen(url)
#raw_data_open = uopen.read()
#uopen.close()

image_open = Image.open(BytesIO(url.content))
my_img = ImageTk.PhotoImage((image_open).resize((640,240)))
my_label = Label(image=my_img)
my_label.grid(row = 0, column = 0,rowspan=3,columnspan=8,sticky=W+E+N+S,padx=5,pady =5)
canvas2 = Canvas(root,width = 640,height = 50, bg = "snow")
canvas2.grid(row = 6, column = 0, columnspan =6,padx = 5, pady = 5)
canvas2.create_text(320,20,font = "Times 14 bold italic",text ="Hussmann")

unlabel1 = Label(root, text = "User Name", bg ="white")
unlabel1.grid(row=7,column =0 ,columnspan = 1,sticky = E,padx = 5, pady = 5)

fromlabel = Label(root,text = "From Date: ", bg = "white")
fromlabel.grid(row=8,column =0 ,columnspan = 1,sticky = E,padx = 5, pady = 5)

pwlabel2 = Label(root,text = "Password", bg = "white")
pwlabel2.grid(row=7,column = 3,columnspan = 1,sticky = E,padx = 5, pady = 5)

tolabel = Label(root,text = "To Date: ", bg = "white")
tolabel.grid(row=8,column =3 ,columnspan = 1,sticky = E,padx = 5, pady = 5)

unentry1 = Entry(root,textvariable = username)
unentry1.grid(row=7,column = 1,columnspan = 1, sticky = W,padx = 5, pady = 5)

entryfrom = Entry(root,textvariable = fromdate)
entryfrom.grid(row=8,column =1 ,columnspan = 1,sticky = W,padx = 5, pady = 5)
fromdate.set(y)

pwentry2 = Entry(root,show = "*",textvariable = userpassword)
pwentry2.grid(row=7,column = 4,columnspan = 1,sticky = W,padx = 5, pady = 5)
userpassword.set("Password")
username.set("USERNAME")

inboxname = Label(root,text="Inbox Folder Name: ex. /Hussmann Folder",bg= "white")
inboxname.grid(row = 13, column = 2,columnspan = 1, sticky = E, padx = 5, pady = 5)
inboxentry = Entry(root,textvariable = inboxexample)
inboxentry.grid(row = 13, column =3, columnspan = 1, sticky = W, padx = 5 , pady = 5)
inboxexample.set("")

def userinfoget():
    n = username.get()
    s = userpassword.get()
    userinbox = inboxexample.get()
    todayscurrentdate = currentdate.get()
    yesterdaysdate = fromdate.get()
    RunProgram(n,s,userinbox,todayscurrentdate,yesterdaysdate)

entryto = Entry(root,textvariable = currentdate)
entryto.grid(row=8,column =4 ,columnspan = 1,sticky = W,padx = 5, pady = 5)
currentdate.set(t)

start_button = Button(root,text = "Run Program", command = userinfoget ,activebackground = "pale green", bg = "white")
start_button.grid(row = 9,column = 2,padx = 5, pady = 5)

button_quit = Button(root, text= "Exit Program", command = root.quit, activebackground= "firebrick1",bg = "white")
button_quit.grid(row=9, column = 5,padx = 5, pady = 5)

root.mainloop()
